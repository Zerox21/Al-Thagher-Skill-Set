from __future__ import annotations
import os, json
from datetime import datetime
from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app
from flask_login import login_required, current_user
from .. import db
from ..models import User, Skill, StudentSkill, Attempt, RemediationUpload, Question
from ..utils import safe_filename

bp = Blueprint("teacher", __name__)

def _ensure_teacher():
    if current_user.role != "teacher":
        flash("Teacher access only.", "error")
        return False
    return True


def _can_unlock_skill_for_student(student_id: str, skill_id: int) -> tuple[bool, str]:
    """To unlock a skill, previous skill must be PASS, or remediation uploaded after last FAIL."""
    skill = Skill.query.get(skill_id)
    if not skill:
        return False, "Skill not found."

    prev = Skill.query.filter(Skill.is_active.is_(True), Skill.order_index == (skill.order_index - 1)).first()
    if not prev:
        return True, ""

    prev_attempt = Attempt.query.filter_by(student_id=student_id, skill_id=prev.id).filter(Attempt.finished_at.isnot(None)).order_by(Attempt.finished_at.desc()).first()
    if not prev_attempt:
        return False, f"Cannot unlock: student hasn't attempted previous skill ({prev.name}) yet."

    if prev_attempt.passed:
        return True, ""

    rem = RemediationUpload.query.filter_by(teacher_id=current_user.id, student_id=student_id, skill_id=prev.id).order_by(RemediationUpload.uploaded_at.desc()).first()
    if rem and rem.uploaded_at and prev_attempt.finished_at and rem.uploaded_at > prev_attempt.finished_at:
        return True, ""

    return False, f"Cannot unlock next skill: previous skill ({prev.name}) is FAIL. Upload remediation for that skill first."


@bp.get("/dashboard")
@login_required
def dashboard():
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    students = User.query.filter_by(role="student", teacher_id=current_user.id).order_by(User.name.asc()).all()
    skills = Skill.query.filter_by(is_active=True).order_by(Skill.order_index.asc()).all()
    attempts = Attempt.query.filter_by(teacher_id=current_user.id).filter(Attempt.finished_at.isnot(None)).all()
    avg_score = round(100 * (sum([a.score or 0 for a in attempts]) / len(attempts)), 2) if attempts else 0.0

    student_rows = []
    for s in students:
        s_attempts = [a for a in attempts if a.student_id == s.id]
        student_rows.append({
            "student": s,
            "attempts": len(s_attempts),
            "avg": round(100 * (sum([a.score or 0 for a in s_attempts]) / len(s_attempts)), 1) if s_attempts else 0
        })
    return render_template("teacher_dashboard.html", students=students, skills=skills, avg_score=avg_score, student_rows=student_rows)

@bp.get("/students/<student_id>")
@login_required
def student_detail(student_id: str):
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    student = User.query.filter_by(id=student_id, role="student", teacher_id=current_user.id).first()
    if not student:
        flash("Student not found.", "error")
        return redirect(url_for("teacher.dashboard"))

    skills = Skill.query.filter_by(is_active=True).order_by(Skill.order_index.asc()).all()
    perms = {p.skill_id: p for p in StudentSkill.query.filter_by(student_id=student.id).all()}
    attempts = Attempt.query.filter_by(student_id=student.id, teacher_id=current_user.id).order_by(Attempt.started_at.desc()).all()
    rem_files = RemediationUpload.query.filter_by(student_id=student.id, teacher_id=current_user.id).order_by(RemediationUpload.uploaded_at.desc()).all()
    return render_template("teacher_student.html", student=student, skills=skills, perms=perms, attempts=attempts, rem_files=rem_files)

@bp.post("/students/<student_id>/toggle_skill")
@login_required
def toggle_skill(student_id: str):
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    student = User.query.filter_by(id=student_id, role="student", teacher_id=current_user.id).first()
    if not student:
        flash("Student not found.", "error")
        return redirect(url_for("teacher.dashboard"))

    skill_id = int(request.form.get("skill_id"))
    allowed = (request.form.get("allowed") == "1")

    if allowed:
        ok, msg = _can_unlock_skill_for_student(student.id, skill_id)
        if not ok:
            flash(msg, "error")
            return redirect(url_for("teacher.student_detail", student_id=student.id))

    perm = StudentSkill.query.filter_by(student_id=student.id, skill_id=skill_id).first()
    if not perm:
        perm = StudentSkill(student_id=student.id, skill_id=skill_id, allowed=allowed, unlocked_at=datetime.utcnow() if allowed else None)
        db.session.add(perm)
    else:
        perm.allowed = allowed
        if allowed and perm.unlocked_at is None:
            perm.unlocked_at = datetime.utcnow()
    db.session.commit()

    flash("Updated skill permission.", "ok")
    return redirect(url_for("teacher.student_detail", student_id=student.id))

@bp.post("/students/<student_id>/upload_remediation")
@login_required
def upload_remediation(student_id: str):
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    student = User.query.filter_by(id=student_id, role="student", teacher_id=current_user.id).first()
    if not student:
        flash("Student not found.", "error")
        return redirect(url_for("teacher.dashboard"))

    skill_id = int(request.form.get("skill_id"))
    note = (request.form.get("note") or "").strip()

    f = request.files.get("file")
    if not f or not f.filename:
        flash("Select a file.", "error")
        return redirect(url_for("teacher.student_detail", student_id=student.id))

    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in current_app.config["ALLOWED_UPLOAD_EXT"]:
        flash("File type not allowed.", "error")
        return redirect(url_for("teacher.student_detail", student_id=student.id))

    safe = safe_filename(f.filename)
    stored_dir = os.path.join(current_app.config['UPLOADS_DIR'], 'teacher', current_user.id, student.id, str(skill_id))
    os.makedirs(stored_dir, exist_ok=True)
    stored_path = os.path.join(stored_dir, safe)
    f.save(stored_path)

    rel = os.path.relpath(stored_path, current_app.config['UPLOADS_DIR'])
    up = RemediationUpload(
        teacher_id=current_user.id,
        student_id=student.id,
        skill_id=skill_id,
        filename=safe,
        stored_path=rel,
        note=note or None
    )
    db.session.add(up)
    db.session.commit()

    flash("Uploaded successfully.", "ok")
    return redirect(url_for("teacher.student_detail", student_id=student.id))


@bp.get("/media")
@login_required
def media_library():
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))
    import os
    from flask import current_app
    media_dir = os.path.join(current_app.config["MEDIA_DIR"], "teacher", current_user.id)
    os.makedirs(media_dir, exist_ok=True)
    files = [n for n in sorted(os.listdir(media_dir)) if os.path.isfile(os.path.join(media_dir, n))]
    return render_template("teacher_media.html", files=files)

@bp.post("/media/upload")
@login_required
def media_upload():
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))
    import os
    from flask import current_app
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Select a file.", "error")
        return redirect(url_for("teacher.media_library"))
    ext = f.filename.rsplit(".",1)[-1].lower() if "." in f.filename else ""
    if ext not in current_app.config["MEDIA_ALLOWED_EXT"]:
        flash("Media type not allowed.", "error")
        return redirect(url_for("teacher.media_library"))

    safe = "".join([c if c.isalnum() or c in "._-" else "_" for c in f.filename]).strip("_") or ("media."+ext)
    media_dir = os.path.join(current_app.config["MEDIA_DIR"], "teacher", current_user.id)
    os.makedirs(media_dir, exist_ok=True)
    f.save(os.path.join(media_dir, safe))

    flash("Uploaded.", "ok")
    return redirect(url_for("teacher.media_library"))

@bp.get("/reports")
@login_required
def reports():
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    attempts = Attempt.query.filter_by(teacher_id=current_user.id).filter(Attempt.finished_at.isnot(None)).order_by(Attempt.finished_at.desc()).limit(200).all()
    return render_template("teacher_reports.html", attempts=attempts)

@bp.get("/download_report/<int:attempt_id>")
@login_required
def download_report(attempt_id: int):
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))
    return redirect(url_for('files.report', attempt_id=attempt_id))

@bp.get("/question_tool")
@login_required
def question_tool():
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    skills = Skill.query.filter_by(is_active=True).order_by(Skill.order_index.asc()).all()
    questions = Question.query.order_by(Question.id.desc()).limit(200).all()
    return render_template("question_tool.html", skills=skills, questions=questions, role=current_user.role)

@bp.post("/question_tool/add")
@login_required
def add_question():
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    skill_id = int(request.form.get("skill_id"))
    qtype = request.form.get("qtype")
    prompt = (request.form.get("prompt") or "").strip()
    options = (request.form.get("options") or "").strip()
    answer = (request.form.get("answer") or "").strip()
    meta = (request.form.get("meta") or "").strip()

    if not prompt:
        flash("Prompt required.", "error")
        return redirect(url_for("teacher.question_tool"))

    # options list
    options_json = None
    if options:
        opts = [x.strip() for x in options.split("\n") if x.strip()]
        options_json = json.dumps(opts, ensure_ascii=False)

    # answer
    answer_json = None
    try:
        if qtype == "mcq_multi":
            answer_json = json.dumps([int(x.strip()) for x in answer.split(",") if x.strip()], ensure_ascii=False)
        elif qtype == "short_text":
            answer_json = json.dumps(answer, ensure_ascii=False)
        else:
            answer_json = json.dumps(int(answer), ensure_ascii=False) if answer != "" else None
    except Exception:
        answer_json = None

    # meta
    meta_json = None
    if meta:
        try:
            meta_json = meta if (meta.strip().startswith("{") or meta.strip().startswith("[")) else json.dumps(meta, ensure_ascii=False)
        except Exception:
            meta_json = None

    q = Question(skill_id=skill_id, qtype=qtype, prompt=prompt, options_json=options_json, answer_json=answer_json, meta_json=meta_json,
               status='draft', created_by_id=current_user.id, created_by_role=current_user.role)
    db.session.add(q)
    db.session.commit()

    flash("Question added.", "ok")
    return redirect(url_for("teacher.question_tool"))

@bp.get("/question_import")
@login_required
def question_import():
    if not _ensure_teacher():
        return redirect(url_for("auth.home"))
    skills = Skill.query.filter_by(is_active=True).order_by(Skill.order_index.asc()).all()
    return render_template("teacher_question_import.html", skills=skills)

@bp.post("/question_import/upload")
@login_required
def question_import_upload():
    if not _ensure_teacher():
        return redirect(url_for("auth.home"))

    import io
    import csv
    import json
    from openpyxl import load_workbook

    f = request.files.get("file")
    default_skill_id = int(request.form.get("default_skill_id") or "0") or None

    if not f or not f.filename:
        flash("Upload a CSV/XLSX file.", "error")
        return redirect(url_for("teacher.question_import"))

    name = f.filename.lower()
    rows = []

    if name.endswith(".csv"):
        content = f.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
    elif name.endswith(".xlsx"):
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
            rows.append(d)
    else:
        flash("Only .csv or .xlsx supported.", "error")
        return redirect(url_for("teacher.question_import"))

    created = 0
    skipped = 0

    for row in rows:
        skill_id = row.get("skill_id")
        skill_name = row.get("skill_name")
        qtype = (row.get("qtype") or "").strip()
        prompt = (row.get("prompt") or "").strip()
        if not prompt or not qtype:
            skipped += 1
            continue

        sid = None
        try:
            sid = int(skill_id) if skill_id not in (None, "", "None") else None
        except Exception:
            sid = None
        if sid is None and default_skill_id:
            sid = default_skill_id
        if sid is None and skill_name:
            sk = Skill.query.filter_by(name=str(skill_name).strip()).first()
            sid = sk.id if sk else None
        if sid is None:
            skipped += 1
            continue

        options_raw = (row.get("options") or "").strip()
        options_json = None
        if options_raw:
            opts = [x.strip() for x in str(options_raw).split("|") if x.strip()]
            options_json = json.dumps(opts, ensure_ascii=False)

        ans_raw = row.get("answer")
        answer_json = None
        try:
            if qtype == "mcq_multi":
                answer_json = json.dumps([int(x.strip()) for x in str(ans_raw).split(",") if x.strip()], ensure_ascii=False)
            elif qtype == "short_text":
                answer_json = json.dumps(str(ans_raw or ""), ensure_ascii=False)
            else:
                answer_json = json.dumps(int(ans_raw), ensure_ascii=False) if ans_raw not in (None, "", "None") else None
        except Exception:
            answer_json = json.dumps(str(ans_raw or ""), ensure_ascii=False) if qtype == "short_text" else None

        meta = (row.get("meta_json") or row.get("meta") or "")
        meta_json = None
        if meta not in (None, "", "None"):
            m = str(meta).strip()
            meta_json = m if (m.startswith("{") or m.startswith("[")) else json.dumps(m, ensure_ascii=False)

        q = Question(skill_id=sid, qtype=qtype, prompt=prompt, options_json=options_json, answer_json=answer_json, meta_json=meta_json,
               status='draft', created_by_id=current_user.id, created_by_role=current_user.role)
        db.session.add(q)
        created += 1

    db.session.commit()
    flash(f"Imported. Created: {created}, Skipped: {skipped}.", "ok")
    return redirect(url_for("teacher.question_tool"))

@bp.get("/export/attempts.csv")
@login_required
def export_attempts_csv():
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    import csv, io
    from flask import Response

    attempts = Attempt.query.filter_by(teacher_id=current_user.id).filter(Attempt.finished_at.isnot(None)).order_by(Attempt.finished_at.desc()).all()
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["attempt_id","student_id","student_name","skill","finished_at","duration_sec","score_pct","passed"])
    for a in attempts:
        w.writerow([a.id, a.student_id, a.student.name if a.student else "", a.skill.name if a.skill else "", a.finished_at, a.duration_sec, int(round((a.score or 0)*100)), a.passed])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition":"attachment; filename=teacher_attempts.csv"})

@bp.get("/export/students.csv")
@login_required
def export_students_csv():
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    import csv, io
    from flask import Response

    students = User.query.filter_by(role="student", teacher_id=current_user.id).order_by(User.id.asc()).all()
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["student_id","name"])
    for s in students:
        w.writerow([s.id, s.name])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition":"attachment; filename=teacher_students.csv"})


@bp.get("/question_tool/edit/<int:question_id>")
@login_required
def question_edit(question_id: int):
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    q = Question.query.get(question_id)
    if not q:
        flash("Question not found.", "error")
        return redirect(url_for("teacher.question_tool"))

    # Teachers can edit only drafts they created (or legacy rows without created_by)
    if q.status == "approved":
        flash("Approved questions are locked (chairman can edit).", "error")
        return redirect(url_for("teacher.question_tool"))
    if q.created_by_id and q.created_by_id != current_user.id and q.created_by_role == "teacher":
        flash("You can only edit your own draft questions.", "error")
        return redirect(url_for("teacher.question_tool"))

    skills = Skill.query.filter_by(is_active=True).order_by(Skill.order_index.asc()).all()
    return render_template("teacher_question_edit.html", q=q, skills=skills)

@bp.post("/question_tool/edit/<int:question_id>")
@login_required
def question_edit_post(question_id: int):
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    import json
    q = Question.query.get(question_id)
    if not q:
        flash("Question not found.", "error")
        return redirect(url_for("teacher.question_tool"))
    if q.status == "approved":
        flash("Approved questions are locked (chairman can edit).", "error")
        return redirect(url_for("teacher.question_tool"))

    q.skill_id = int(request.form.get("skill_id"))
    q.qtype = request.form.get("qtype")
    q.prompt = (request.form.get("prompt") or "").strip()

    options = (request.form.get("options") or "").strip()
    answer = (request.form.get("answer") or "").strip()
    meta = (request.form.get("meta") or "").strip()

    if options:
        opts = [x.strip() for x in options.split("\n") if x.strip()]
        q.options_json = json.dumps(opts, ensure_ascii=False)
    else:
        q.options_json = None

    try:
        if q.qtype == "mcq_multi":
            q.answer_json = json.dumps([int(x.strip()) for x in answer.split(",") if x.strip()], ensure_ascii=False)
        elif q.qtype == "short_text":
            q.answer_json = json.dumps(answer, ensure_ascii=False)
        else:
            q.answer_json = json.dumps(int(answer), ensure_ascii=False) if answer != "" else None
    except Exception:
        q.answer_json = json.dumps(answer, ensure_ascii=False) if q.qtype == "short_text" else None

    if meta:
        m = meta.strip()
        q.meta_json = m if (m.startswith("{") or m.startswith("[")) else json.dumps(m, ensure_ascii=False)
    else:
        q.meta_json = None

    db.session.commit()
    flash("Draft updated.", "ok")
    return redirect(url_for("teacher.question_tool"))

@bp.get("/doc_import")
@login_required
def doc_import():
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))
    skills = Skill.query.filter_by(is_active=True).order_by(Skill.order_index.asc()).all()
    return render_template("teacher_doc_import.html", skills=skills)

@bp.post("/doc_import")
@login_required
def doc_import_post():
    if not _ensure_teacher():
        return redirect(url_for('auth.home'))

    import os, json
    from flask import current_app
    from ..doc_import import render_pdf_pages_to_images, split_docx_text_to_questions
    from docx import Document

    f = request.files.get("file")
    default_skill_id = int(request.form.get("default_skill_id") or "0") or None
    mode = request.form.get("mode") or "auto"

    if not f or not f.filename:
        flash("Upload a PDF or DOCX.", "error")
        return redirect(url_for("teacher.doc_import"))

    name = f.filename
    ext = name.rsplit(".",1)[-1].lower() if "." in name else ""
    if ext not in {"pdf","docx"}:
        flash("Only PDF or DOCX supported.", "error")
        return redirect(url_for("teacher.doc_import"))

    if not default_skill_id:
        flash("Choose a default skill.", "error")
        return redirect(url_for("teacher.doc_import"))

    # store original under MEDIA_DIR/imports/teacher/<id>/
    base_dir = os.path.join(current_app.config["MEDIA_DIR"], "imports", "teacher", current_user.id)
    os.makedirs(base_dir, exist_ok=True)
    safe = "".join([c if c.isalnum() or c in "._-" else "_" for c in name]).strip("_") or f"upload.{ext}"
    src_path = os.path.join(base_dir, safe)
    f.save(src_path)

    created = 0
    if ext == "pdf":
        out_dir = os.path.join(current_app.config["MEDIA_DIR"], "teacher", current_user.id, "pdf_imports")
        prefix =_toggle_prefix(safe)
        imgs = render_pdf_pages_to_images(src_path, out_dir, prefix=prefix)
        for img in imgs:
            rel = f"teacher/{current_user.id}/pdf_imports/{img}"
            q = Question(
                skill_id=default_skill_id,
                qtype="image_mcq_single",
                prompt=f"Imported from PDF: {safe} — page {img.split('_p')[-1].split('.')[0]}",
                options_json=None,
                answer_json=None,
                meta_json=json.dumps({"image_media": rel}, ensure_ascii=False),
                status="draft",
                created_by_id=current_user.id,
                created_by_role="teacher",
            )
            db.session.add(q)
            created += 1
        db.session.commit()
        flash(f"Imported {created} draft questions (one per page). Edit each to add options & answer.", "ok")
        return redirect(url_for("teacher.question_tool"))

    # DOCX
    doc = Document(src_path)
    full_text = "\n".join([p.text for p in doc.paragraphs if p.text])
    qs = split_docx_text_to_questions(full_text)
    if not qs:
        qs = [full_text] if full_text.strip() else []
    for i, qtext in enumerate(qs, start=1):
        q = Question(
            skill_id=default_skill_id,
            qtype="short_text",
            prompt=f"Imported from DOCX: {safe}\n\n{qtext}",
            options_json=None,
            answer_json=None,
            meta_json=None,
            status="draft",
            created_by_id=current_user.id,
            created_by_role="teacher",
        )
        db.session.add(q)
        created += 1
    db.session.commit()
    flash(f"Imported {created} draft questions from DOCX. Review and set type/options/answer.", "ok")
    return redirect(url_for("teacher.question_tool"))

def _toggle_prefix(filename: str) -> str:
    base = filename.rsplit(".",1)[0]
    base = re.sub(r"[^A-Za-z0-9_-]+", "_", base)
    return base[:40] or "pdf"
